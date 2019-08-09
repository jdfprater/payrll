#! python3

import openpyxl, json, re
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


adpHoursWB = openpyxl.load_workbook('Employee Transactions and Totals (Excel).xlsx')
hsheet = adpHoursWB['Sheet1']


def noMiddle(name):
    #find any number of any character, comma, space, any number of any character, space, one of any letter 
    middleRegex = re.compile(r'\D*,\s\D*\s\D{1}')
    mo = middleRegex.search(name)
    if mo:
        name = name[:-2]
    return name
    
employees = {}


row_count = hsheet.max_row    
name = column_index_from_string('A')
hours = column_index_from_string('G')


for row in hsheet.rows:
    emp_name = noMiddle(str(row[name-1].value))
    emp_hours = row[hours-1].value
    if emp_hours is None:
        continue
    else:
        if emp_name not in employees and ',' in emp_name:
            emp_hours = float(emp_hours)
            employees.update( { emp_name : {'hours' : emp_hours, 'FP' : 0, 'sessions' : 0.0, 'classes' : 0}})
            #print(emp_name, employees[emp_name])
        elif emp_name in employees:
            emp_hours = float(emp_hours)
            employees[emp_name]['hours'] += emp_hours
            #print('adding to ' + emp_name, employees[emp_name])


#print(json.dumps(employees, indent=1))



#open pt training payroll report
ptTrainingWB = openpyxl.load_workbook('PT Training Payroll Report.xlsx') #open 'pt training payroll report.xlsx'
tsheet = ptTrainingWB['PT_Payroll_Detail'] #select sheet


location_club = column_index_from_string('J')
bonusHours = column_index_from_string('P')
trainer = column_index_from_string('E')
agreement = column_index_from_string('N')
fp_type = ['GOLD\'S 3D', 'Fitness Profile', 'Fitness Profile Follow-Up', 'Fit Profile', 'Fit Profile Follow-Up', 'Fitness Assessment']

for row in tsheet.rows:
    pt_name = row[trainer-1].value
    gym_name = row[location_club-1].value
    hours = row[bonusHours-1].value
    session_type = row[agreement-1].value
    if pt_name in employees:
        #employees[pt_name]['gym'] = gym_name
        if any(x in (session_type) for x in fp_type):
            employees[pt_name]['FP'] += 1
        elif session_type != 'GGX Group Trackers' and hours > 0:
            employees[pt_name]['sessions'] += hours

#open service provider scheduler report
classesWB = openpyxl.load_workbook('Daily Service Provider Scheduler.xlsx')
csheet = classesWB['Sheet1']

club_name = column_index_from_string('C')
service_provider = column_index_from_string('A')
event = column_index_from_string('U')
attendance = column_index_from_string('V')
studio = ['BOOTCAMP', 'GOLD\'S FIT', 'GOLD\'S CYCLE', 'GOLD\'S CYCLE BEATS', 'GOLD\'S CYCLE', 'STUDIO FUSION', 'GOLD\'S BURN']

for row in csheet.rows:
    gym_name = row[club_name-1].value
    instructor = row[service_provider-1].value
    classes = row[event-1].value
    attendees = row[attendance-1].value
    if instructor in employees and classes:
        #employees[instructor]['gym'] = gym_name
        if any(x in (classes) for x in studio) and attendees > 0:
            employees[instructor]['classes'] += 1.5
        else:
            employees[instructor]['classes'] += 1
    else:
        continue


#open a spreadsheet to write in
testpayrollWB = Workbook()
psheet = testpayrollWB.active
#               A-1                 B-2             C-3         D-4         E-5                 F-6
headers = ['employee name', 'hours clocked-in', 'FP show', 'sessions', 'classes scheduled', 'hours over/under']

for i in range(len(headers)):
    psheet.cell(row=1, column=i+1).font = Font(bold=True)
    psheet.cell(row=1, column=i+1).value = headers[i]

row = 2
for employee, work_hours in employees.items():
    psheet.cell(row=row, column=1, value=employee)
    column = 2
    for type_hours, num_hours in work_hours.items():
        num_hours = float(num_hours)
        psheet.cell(row=row, column=column, value=num_hours)
        column += 1
    row += 1


#calculate remaining hours
for i in range(len(employees)):
    psheet.cell(row=i+2, column =6).value = '=B' + str(i+2) + '-(C' + str(i+2) + '+D' + str(i+2) + '+E' + str(i+2) + ')'
    if int(psheet.cell(row=i+2, column=2).value - (psheet.cell(row=i+2, column=3).value + psheet.cell(row=i+2, column=4).value + psheet.cell(row=i+2, column=5).value)) > 2:
        psheet.cell(row=i+2, column=6).fill = PatternFill(fgColor='FFC7CE', fill_type = 'solid')



testpayrollWB.save('payroll_test.xlsx')
testpayrollWB.close()
